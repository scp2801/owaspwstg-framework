"""
Professional Report Engine
==========================
Generates comprehensive security reports in multiple formats:
- HTML (Professional pentest report)
- Excel (OWASP WSTG Checklist - Professional)
- JSON (Machine-readable findings)
- Markdown (GitHub-ready)
- TXT (Plain text summary)

Excel report follows OWASP WSTG checklist format exactly.
"""

import os
import json
from datetime import datetime
from typing import Dict, List, Any
from pathlib import Path

from core.utils import sanitize_filename


class ReportEngine:
    """
    Generates professional security reports in multiple formats.
    """

    def __init__(self, config: Dict, logger):
        self.config = config
        self.logger = logger

    def generate_all(self, results: Dict, output_dir: str, formats: str = "all") -> Dict[str, str]:
        """
        Generate reports in all requested formats.

        Args:
            results: Complete scan results dictionary
            output_dir: Output directory path
            formats: Comma-separated format list or 'all'

        Returns:
            Dictionary mapping format name to output file path
        """
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        report_paths = {}

        requested = [f.strip().lower() for f in formats.split(",")] if formats != "all" else \
            ["html", "json", "excel", "markdown", "txt"]

        if formats == "all":
            requested = ["html", "json", "excel", "markdown", "txt"]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = results.get("targets", ["unknown"])[0]
        base_name = sanitize_filename(target) + f"_{timestamp}"
        scan_dir = results.get("scan_dir", output_dir)

        for fmt in requested:
            try:
                if fmt == "html":
                    path = os.path.join(scan_dir, f"{base_name}_report.html")
                    self._generate_html(results, path)
                    report_paths["html"] = path

                elif fmt == "json":
                    path = os.path.join(scan_dir, f"{base_name}_report.json")
                    self._generate_json(results, path)
                    report_paths["json"] = path

                elif fmt in ("excel", "xlsx"):
                    path = os.path.join(scan_dir, f"{base_name}_WSTG_checklist.xlsx")
                    self._generate_excel(results, path)
                    report_paths["excel"] = path

                elif fmt in ("markdown", "md"):
                    path = os.path.join(scan_dir, f"{base_name}_report.md")
                    self._generate_markdown(results, path)
                    report_paths["markdown"] = path

                elif fmt == "txt":
                    path = os.path.join(scan_dir, f"{base_name}_report.txt")
                    self._generate_txt(results, path)
                    report_paths["txt"] = path

            except Exception as e:
                self.logger.error(f"Report generation failed for {fmt}: {e}")

        return report_paths

    # ──────────────────────────────────────────────────────────────────
    # HTML REPORT
    # ──────────────────────────────────────────────────────────────────
    def _generate_html(self, results: Dict, path: str):
        """Generate professional HTML pentest report."""
        target = results.get("targets", ["Unknown"])[0]
        scan_id = results.get("scan_id", "N/A")
        scan_date = results.get("scan_start", datetime.now().isoformat())[:10]
        checklist = results.get("checklist", [])
        checklist_summary = results.get("checklist_summary", {})
        targets_results = results.get("targets_results", [{}])
        tr = targets_results[0] if targets_results else {}

        vulns = tr.get("vulnerabilities", [])
        critical = [v for v in vulns if v.get("severity") == "Critical"]
        high = [v for v in vulns if v.get("severity") == "High"]
        medium = [v for v in vulns if v.get("severity") == "Medium"]
        low = [v for v in vulns if v.get("severity") == "Low"]

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>OWASP WSTG Security Assessment - {target}</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #0a0e1a; color: #e0e6f0; line-height: 1.6; }}
  .header {{ background: linear-gradient(135deg, #0d1b2a 0%, #1a2840 50%, #0d1b2a 100%); padding: 40px; border-bottom: 3px solid #00d4aa; }}
  .header h1 {{ font-size: 2.2em; color: #00d4aa; font-weight: 700; }}
  .header .subtitle {{ color: #a0b4c8; margin-top: 8px; font-size: 1.1em; }}
  .meta-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 20px; padding: 30px 40px; background: #0f1623; }}
  .meta-card {{ background: #1a2535; border-radius: 8px; padding: 20px; border-left: 4px solid #00d4aa; }}
  .meta-card .label {{ color: #7a8fa6; font-size: 0.85em; text-transform: uppercase; letter-spacing: 1px; }}
  .meta-card .value {{ color: #ffffff; font-size: 1.4em; font-weight: 700; margin-top: 5px; }}
  .section {{ padding: 30px 40px; }}
  .section-title {{ font-size: 1.4em; color: #00d4aa; border-bottom: 2px solid #1e3040; padding-bottom: 10px; margin-bottom: 20px; }}
  .severity-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 30px; }}
  .sev-card {{ padding: 20px; border-radius: 8px; text-align: center; }}
  .sev-card.critical {{ background: rgba(220, 38, 38, 0.15); border: 2px solid #dc2626; }}
  .sev-card.high {{ background: rgba(234, 88, 12, 0.15); border: 2px solid #ea580c; }}
  .sev-card.medium {{ background: rgba(234, 179, 8, 0.15); border: 2px solid #eab308; }}
  .sev-card.low {{ background: rgba(59, 130, 246, 0.15); border: 2px solid #3b82f6; }}
  .sev-card .count {{ font-size: 2.5em; font-weight: 700; }}
  .sev-card .label {{ font-size: 0.9em; text-transform: uppercase; letter-spacing: 1px; margin-top: 5px; }}
  .critical .count {{ color: #ef4444; }} .critical .label {{ color: #dc2626; }}
  .high .count {{ color: #f97316; }} .high .label {{ color: #ea580c; }}
  .medium .count {{ color: #fbbf24; }} .medium .label {{ color: #eab308; }}
  .low .count {{ color: #60a5fa; }} .low .label {{ color: #3b82f6; }}
  .finding {{ background: #1a2535; border-radius: 8px; padding: 20px; margin-bottom: 15px; border-left: 4px solid #333; }}
  .finding.critical {{ border-left-color: #ef4444; }}
  .finding.high {{ border-left-color: #f97316; }}
  .finding.medium {{ border-left-color: #fbbf24; }}
  .finding.low {{ border-left-color: #60a5fa; }}
  .finding.info {{ border-left-color: #22d3ee; }}
  .finding-header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px; }}
  .finding-title {{ font-size: 1.1em; font-weight: 600; color: #ffffff; }}
  .badge {{ padding: 4px 12px; border-radius: 20px; font-size: 0.8em; font-weight: 600; text-transform: uppercase; }}
  .badge.critical {{ background: #7f1d1d; color: #fca5a5; }}
  .badge.high {{ background: #7c2d12; color: #fdba74; }}
  .badge.medium {{ background: #713f12; color: #fde68a; }}
  .badge.low {{ background: #1e3a5f; color: #93c5fd; }}
  .badge.info {{ background: #164e63; color: #67e8f9; }}
  .finding-meta {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin: 10px 0; }}
  .finding-field label {{ color: #7a8fa6; font-size: 0.85em; }}
  .finding-field span {{ color: #c0cfe0; font-size: 0.9em; }}
  .evidence-box {{ background: #0a0e1a; border-radius: 6px; padding: 12px; margin-top: 10px; font-family: 'Courier New', monospace; font-size: 0.85em; color: #7dd3fc; white-space: pre-wrap; word-break: break-all; max-height: 150px; overflow-y: auto; }}
  .rec-box {{ background: #0f2718; border-radius: 6px; padding: 12px; margin-top: 8px; color: #86efac; font-size: 0.9em; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
  th {{ background: #0f2030; color: #00d4aa; padding: 10px; text-align: left; font-size: 0.85em; text-transform: uppercase; }}
  td {{ padding: 10px; border-bottom: 1px solid #1e3040; font-size: 0.9em; }}
  tr:hover {{ background: rgba(0, 212, 170, 0.05); }}
  .status-pass {{ color: #4ade80; font-weight: 600; }}
  .status-fail {{ color: #f87171; font-weight: 600; }}
  .status-review {{ color: #fbbf24; font-weight: 600; }}
  .status-na {{ color: #6b7280; }}
  .footer {{ background: #0d1b2a; border-top: 2px solid #1e3040; padding: 20px 40px; color: #4a5568; font-size: 0.85em; text-align: center; }}
  .disclaimer {{ background: rgba(234, 88, 12, 0.1); border: 1px solid #ea580c; border-radius: 8px; padding: 15px; margin: 20px 40px; color: #fdba74; }}
</style>
</head>
<body>

<div class="header">
  <h1>🔒 Security Assessment Report</h1>
  <div class="subtitle">OWASP Web Security Testing Guide (WSTG) v4.2 — Automated Analysis</div>
</div>

<div class="meta-grid">
  <div class="meta-card">
    <div class="label">Target</div>
    <div class="value" style="font-size:1.1em">{target}</div>
  </div>
  <div class="meta-card">
    <div class="label">Scan Date</div>
    <div class="value" style="font-size:1.1em">{scan_date}</div>
  </div>
  <div class="meta-card">
    <div class="label">Scan ID</div>
    <div class="value" style="font-size:1em">{scan_id}</div>
  </div>
  <div class="meta-card">
    <div class="label">Duration</div>
    <div class="value" style="font-size:1.1em">{results.get('duration_seconds', 0)}s</div>
  </div>
</div>

<div class="disclaimer">
  ⚠️ <strong>Legal Disclaimer:</strong> This report is for authorized security testing only. 
  Unauthorized use is prohibited. All findings should be responsibly disclosed.
</div>

<div class="section">
  <div class="section-title">📊 Executive Summary</div>
  <div class="severity-grid">
    <div class="sev-card critical">
      <div class="count">{len(critical)}</div>
      <div class="label">Critical</div>
    </div>
    <div class="sev-card high">
      <div class="count">{len(high)}</div>
      <div class="label">High</div>
    </div>
    <div class="sev-card medium">
      <div class="count">{len(medium)}</div>
      <div class="label">Medium</div>
    </div>
    <div class="sev-card low">
      <div class="count">{len(low)}</div>
      <div class="label">Low</div>
    </div>
  </div>

  <table>
    <tr><th>Metric</th><th>Value</th></tr>
    <tr><td>Subdomains Discovered</td><td>{len(tr.get('subdomains', []))}</td></tr>
    <tr><td>Live Hosts</td><td>{len(tr.get('live_hosts', []))}</td></tr>
    <tr><td>Endpoints Crawled</td><td>{len(tr.get('endpoints', []))}</td></tr>
    <tr><td>Parameters Found</td><td>{len(tr.get('parameters', []))}</td></tr>
    <tr><td>JS Files Analyzed</td><td>{len(tr.get('js_files', []))}</td></tr>
    <tr><td>Total Vulnerabilities</td><td>{len(vulns)}</td></tr>
    <tr><td>WSTG Tests Executed</td><td>{checklist_summary.get('total', 0)}</td></tr>
    <tr><td>WSTG Pass Rate</td><td>{checklist_summary.get('pass_rate', 0)}%</td></tr>
  </table>
</div>

<div class="section">
  <div class="section-title">🐛 Vulnerability Findings</div>
"""

        # Add vulnerability findings
        real_vulns = [v for v in vulns if v.get("vulnerable", False)]
        if real_vulns:
            for v in sorted(real_vulns, key=lambda x: {"Critical": 0, "High": 1, "Medium": 2, "Low": 3, "Info": 4}.get(x.get("severity", "Info"), 4)):
                sev = v.get("severity", "info").lower()
                html += f"""
  <div class="finding {sev}">
    <div class="finding-header">
      <span class="finding-title">{v.get('name', 'Unknown Finding')}</span>
      <span class="badge {sev}">{v.get('severity', 'Info')}</span>
    </div>
    <div class="finding-meta">
      <div class="finding-field"><label>WSTG ID: </label><span>{v.get('wstg_id', 'N/A')}</span></div>
      <div class="finding-field"><label>CWE: </label><span>{v.get('cwe', 'N/A')}</span></div>
      <div class="finding-field"><label>CVSS: </label><span>{v.get('cvss', 'N/A')}</span></div>
      <div class="finding-field"><label>Status: </label><span class="status-fail">FAIL</span></div>
    </div>
    <div><strong>URL:</strong> <span style="color:#7dd3fc;word-break:break-all">{v.get('url', 'N/A')}</span></div>
    <div style="margin-top:8px"><strong>Description:</strong> {v.get('description', '')}</div>
    <div class="evidence-box">{v.get('evidence', 'No evidence collected')}</div>
    <div class="rec-box">💡 <strong>Recommendation:</strong> {v.get('recommendation', '')}</div>
  </div>"""
        else:
            html += "<p style='color:#4ade80'>✓ No exploitable vulnerabilities detected.</p>"

        # WSTG Checklist Table
        sc = checklist_summary.get("status_counts", {})
        html += f"""
</div>

<div class="section">
  <div class="section-title">📋 OWASP WSTG Checklist Summary</div>
  <div style="display:grid;grid-template-columns:repeat(4,1fr);gap:15px;margin-bottom:20px">
    <div style="text-align:center;padding:15px;background:#1a2535;border-radius:8px">
      <div style="font-size:2em;color:#4ade80;font-weight:700">{sc.get('PASS', 0)}</div>
      <div style="color:#6b7280;font-size:0.9em">PASS</div>
    </div>
    <div style="text-align:center;padding:15px;background:#1a2535;border-radius:8px">
      <div style="font-size:2em;color:#f87171;font-weight:700">{sc.get('FAIL', 0)}</div>
      <div style="color:#6b7280;font-size:0.9em">FAIL</div>
    </div>
    <div style="text-align:center;padding:15px;background:#1a2535;border-radius:8px">
      <div style="font-size:2em;color:#fbbf24;font-weight:700">{sc.get('REVIEW', 0)}</div>
      <div style="color:#6b7280;font-size:0.9em">REVIEW</div>
    </div>
    <div style="text-align:center;padding:15px;background:#1a2535;border-radius:8px">
      <div style="font-size:2em;color:#6b7280;font-weight:700">{sc.get('NOT_TESTED', 0)}</div>
      <div style="color:#6b7280;font-size:0.9em">NOT TESTED</div>
    </div>
  </div>

  <table>
    <tr>
      <th>WSTG ID</th>
      <th>Category</th>
      <th>Test Name</th>
      <th>Status</th>
      <th>Severity</th>
    </tr>
"""
        for item in checklist:
            status = item.get("status", "NOT_TESTED")
            status_val = status.value if hasattr(status, "value") else str(status)
            css_class = {"PASS": "status-pass", "FAIL": "status-fail", "REVIEW": "status-review"}.get(status_val, "status-na")
            html += f"""
    <tr>
      <td><strong>{item.get('id', '')}</strong></td>
      <td>{item.get('category', '')}</td>
      <td>{item.get('name', '')}</td>
      <td class="{css_class}">{status_val}</td>
      <td>{item.get('severity', '')}</td>
    </tr>"""

        html += """
  </table>
</div>

<div class="footer">
  <p>Generated by <strong>OWASP WSTG Bug Bounty Framework</strong> | github.com/scp2801/owaspwstg-framework</p>
  <p style="margin-top:5px">⚠️ For authorized security testing only. All findings must be responsibly disclosed.</p>
</div>
</body>
</html>"""

        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        self.logger.info(f"HTML report generated: {path}")

    # ──────────────────────────────────────────────────────────────────
    # EXCEL REPORT (Professional OWASP WSTG Checklist)
    # ──────────────────────────────────────────────────────────────────
    def _generate_excel(self, results: Dict, path: str):
        """Generate professional Excel OWASP WSTG checklist report."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                PatternFill, Font, Alignment, Border, Side, GradientFill
            )
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.filters import AutoFilter
        except ImportError:
            self.logger.error("openpyxl not installed. Cannot generate Excel report.")
            return

        wb = Workbook()

        # ── Sheet 1: WSTG Checklist ─────────────────────────────────
        ws = wb.active
        ws.title = "WSTG Checklist"

        # Color definitions
        colors = {
            "header_bg": "1A2840",
            "header_fg": "00D4AA",
            "critical": "7F1D1D",
            "critical_fg": "FCA5A5",
            "high": "7C2D12",
            "high_fg": "FDBA74",
            "medium": "713F12",
            "medium_fg": "FDE68A",
            "low": "1E3A5F",
            "low_fg": "93C5FD",
            "info": "164E63",
            "info_fg": "67E8F9",
            "pass_bg": "14532D",
            "pass_fg": "4ADE80",
            "fail_bg": "7F1D1D",
            "fail_fg": "F87171",
            "review_bg": "713F12",
            "review_fg": "FDE68A",
            "not_tested": "1F2937",
            "row_even": "111827",
            "row_odd": "1F2937",
        }

        def make_fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def make_font(hex_color, bold=False, size=10):
            return Font(color=hex_color, bold=bold, size=size, name="Calibri")

        def make_border():
            side = Side(style="thin", color="2D3748")
            return Border(left=side, right=side, top=side, bottom=side)

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Title row
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = "OWASP Web Security Testing Guide (WSTG) v4.2 — Security Assessment Checklist"
        title_cell.fill = make_fill(colors["header_bg"])
        title_cell.font = Font(color=colors["header_fg"], bold=True, size=14, name="Calibri")
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 35

        # Target info row
        ws.merge_cells("A2:I2")
        target = results.get("targets", ["Unknown"])[0]
        scan_date = results.get("scan_start", "")[:10]
        ws["A2"].value = f"Target: {target}  |  Date: {scan_date}  |  Framework: OWASP WSTG Bug Bounty Framework (scp2801)"
        ws["A2"].fill = make_fill("0F1623")
        ws["A2"].font = Font(color="7A8FA6", size=10, name="Calibri")
        ws["A2"].alignment = center_align
        ws.row_dimensions[2].height = 20

        # Header row
        headers = ["WSTG ID", "Category", "Test Name", "Status", "Severity", "URL", "Evidence", "Recommendation", "Notes"]
        col_widths = [18, 24, 40, 14, 12, 45, 50, 50, 30]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = make_fill("0A1628")
            cell.font = Font(color="00D4AA", bold=True, size=10, name="Calibri")
            cell.alignment = center_align
            cell.border = make_border()
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[3].height = 22

        # Status color mapping
        status_colors = {
            "PASS":       (colors["pass_bg"], colors["pass_fg"]),
            "FAIL":       (colors["fail_bg"], colors["fail_fg"]),
            "REVIEW":     (colors["review_bg"], colors["review_fg"]),
            "NOT_TESTED": (colors["not_tested"], "6B7280"),
            "N/A":        ("374151", "6B7280"),
        }

        severity_colors = {
            "Critical": (colors["critical"], colors["critical_fg"]),
            "High":     (colors["high"], colors["high_fg"]),
            "Medium":   (colors["medium"], colors["medium_fg"]),
            "Low":      (colors["low"], colors["low_fg"]),
            "Info":     (colors["info"], colors["info_fg"]),
        }

        checklist = results.get("checklist", [])

        for row_idx, item in enumerate(checklist, start=4):
            is_even = row_idx % 2 == 0
            row_bg = colors["row_even"] if is_even else colors["row_odd"]

            status = item.get("status", "NOT_TESTED")
            status_val = status.value if hasattr(status, "value") else str(status)
            severity = item.get("severity", "Info")

            # Build URL string
            urls = item.get("urls", [])
            url_str = urls[0] if urls else ""

            # Build evidence string
            evidence_list = item.get("evidence", [])
            evidence_str = "\n".join(evidence_list[:2]) if evidence_list else ""

            row_data = [
                item.get("id", ""),
                item.get("category", ""),
                item.get("name", ""),
                status_val,
                severity,
                url_str[:100] if url_str else "",
                evidence_str[:300] if evidence_str else "",
                item.get("recommendations", "")[:200],
                item.get("notes", "")[:100],
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = make_border()

                if col_idx == 1:  # WSTG ID - always themed
                    cell.fill = make_fill("0A1628")
                    cell.font = Font(color="00D4AA", bold=True, size=9, name="Calibri")
                    cell.alignment = center_align
                elif col_idx == 4:  # Status column
                    bg, fg = status_colors.get(status_val, (row_bg, "FFFFFF"))
                    cell.fill = make_fill(bg)
                    cell.font = Font(color=fg, bold=True, size=9, name="Calibri")
                    cell.alignment = center_align
                elif col_idx == 5:  # Severity column
                    bg, fg = severity_colors.get(severity, (row_bg, "FFFFFF"))
                    cell.fill = make_fill(bg)
                    cell.font = Font(color=fg, bold=True, size=9, name="Calibri")
                    cell.alignment = center_align
                else:
                    cell.fill = make_fill(row_bg)
                    cell.font = Font(color="C0CFE0", size=9, name="Calibri")
                    cell.alignment = left_align if col_idx > 3 else center_align

            ws.row_dimensions[row_idx].height = 30

        # Add AutoFilter
        ws.auto_filter.ref = f"A3:I{3 + len(checklist)}"

        # Freeze panes (keep header visible)
        ws.freeze_panes = "A4"

        # ── Sheet 2: Vulnerability Findings ─────────────────────────
        ws2 = wb.create_sheet(title="Vulnerability Findings")

        ws2.merge_cells("A1:G1")
        ws2["A1"].value = "Security Vulnerability Findings"
        ws2["A1"].fill = make_fill("1A2840")
        ws2["A1"].font = Font(color="00D4AA", bold=True, size=13, name="Calibri")
        ws2["A1"].alignment = center_align
        ws2.row_dimensions[1].height = 30

        vuln_headers = ["Finding Name", "Severity", "WSTG ID", "CWE", "CVSS", "URL", "Recommendation"]
        vuln_widths =  [40, 12, 18, 12, 8, 60, 60]
        for ci, (h, w) in enumerate(zip(vuln_headers, vuln_widths), 1):
            cell = ws2.cell(row=2, column=ci, value=h)
            cell.fill = make_fill("0A1628")
            cell.font = Font(color="00D4AA", bold=True, size=10, name="Calibri")
            cell.alignment = center_align
            cell.border = make_border()
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.row_dimensions[2].height = 22

        all_vulns = []
        for tr_data in results.get("targets_results", []):
            all_vulns.extend(tr_data.get("vulnerabilities", []))

        real_vulns = [v for v in all_vulns if v.get("vulnerable", False)]
        real_vulns.sort(key=lambda x: {"Critical": 0, "High": 1, "Medium": 2, "Low": 3, "Info": 4}.get(x.get("severity", "Info"), 4))

        for ri, v in enumerate(real_vulns, start=3):
            is_even = ri % 2 == 0
            row_bg = colors["row_even"] if is_even else colors["row_odd"]
            severity = v.get("severity", "Info")
            sev_bg, sev_fg = severity_colors.get(severity, (row_bg, "FFFFFF"))

            row_data = [
                v.get("name", ""),
                severity,
                v.get("wstg_id", ""),
                v.get("cwe", ""),
                v.get("cvss", ""),
                v.get("url", "")[:150],
                v.get("recommendation", "")[:200],
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.border = make_border()
                if ci == 2:
                    cell.fill = make_fill(sev_bg)
                    cell.font = Font(color=sev_fg, bold=True, size=9, name="Calibri")
                    cell.alignment = center_align
                else:
                    cell.fill = make_fill(row_bg)
                    cell.font = Font(color="C0CFE0", size=9, name="Calibri")
                    cell.alignment = left_align
            ws2.row_dimensions[ri].height = 28

        ws2.auto_filter.ref = f"A2:G{2 + len(real_vulns)}"
        ws2.freeze_panes = "A3"

        # ── Sheet 3: Recon Summary ───────────────────────────────────
        ws3 = wb.create_sheet(title="Recon Summary")
        ws3.merge_cells("A1:D1")
        ws3["A1"].value = "Reconnaissance Summary"
        ws3["A1"].fill = make_fill("1A2840")
        ws3["A1"].font = Font(color="00D4AA", bold=True, size=13, name="Calibri")
        ws3["A1"].alignment = center_align
        ws3.row_dimensions[1].height = 30

        tr_data = results.get("targets_results", [{}])[0] if results.get("targets_results") else {}

        recon_headers = ["Type", "Value", "Extra Info", "Status"]
        recon_widths = [20, 50, 40, 15]
        for ci, (h, w) in enumerate(zip(recon_headers, recon_widths), 1):
            cell = ws3.cell(row=2, column=ci, value=h)
            cell.fill = make_fill("0A1628")
            cell.font = Font(color="00D4AA", bold=True, size=10, name="Calibri")
            cell.alignment = center_align
            cell.border = make_border()
            ws3.column_dimensions[get_column_letter(ci)].width = w
        ws3.row_dimensions[2].height = 22

        row_num = 3
        # Subdomains
        for sub in tr_data.get("subdomains", [])[:100]:
            ws3.cell(row=row_num, column=1, value="Subdomain").fill = make_fill("111827")
            ws3.cell(row=row_num, column=1).font = Font(color="7A8FA6", size=9, name="Calibri")
            ws3.cell(row=row_num, column=2, value=sub).fill = make_fill("111827")
            ws3.cell(row=row_num, column=2).font = Font(color="C0CFE0", size=9, name="Calibri")
            ws3.cell(row=row_num, column=3, value="").fill = make_fill("111827")
            ws3.cell(row=row_num, column=4, value="Discovered").fill = make_fill("111827")
            ws3.cell(row=row_num, column=4).font = Font(color="4ADE80", size=9, name="Calibri")
            for ci in range(1, 5):
                ws3.cell(row=row_num, column=ci).border = make_border()
            row_num += 1

        # Live hosts
        for host in tr_data.get("live_hosts", [])[:50]:
            ws3.cell(row=row_num, column=1, value="Live Host").fill = make_fill("1F2937")
            ws3.cell(row=row_num, column=1).font = Font(color="7A8FA6", size=9, name="Calibri")
            ws3.cell(row=row_num, column=2, value=host.get("url", "")).fill = make_fill("1F2937")
            ws3.cell(row=row_num, column=2).font = Font(color="C0CFE0", size=9, name="Calibri")
            ws3.cell(row=row_num, column=3, value=f"{host.get('title', '')} | {host.get('server', '')}").fill = make_fill("1F2937")
            ws3.cell(row=row_num, column=3).font = Font(color="7A8FA6", size=9, name="Calibri")
            ws3.cell(row=row_num, column=4, value=f"HTTP {host.get('status_code', '')}").fill = make_fill("1F2937")
            ws3.cell(row=row_num, column=4).font = Font(color="4ADE80", size=9, name="Calibri")
            for ci in range(1, 5):
                ws3.cell(row=row_num, column=ci).border = make_border()
            row_num += 1

        ws3.freeze_panes = "A3"

        wb.save(path)
        self.logger.info(f"Excel report generated: {path}")

    # ──────────────────────────────────────────────────────────────────
    # JSON REPORT
    # ──────────────────────────────────────────────────────────────────
    def _generate_json(self, results: Dict, path: str):
        """Generate machine-readable JSON report."""
        target = results.get("targets", ["Unknown"])[0]
        tr_data = results.get("targets_results", [{}])[0] if results.get("targets_results") else {}

        report = {
            "meta": {
                "framework": "OWASP WSTG Bug Bounty Framework",
                "version": "1.0.0",
                "author": "scp2801",
                "github": "https://github.com/scp2801/owaspwstg-framework",
                "scan_id": results.get("scan_id"),
                "target": target,
                "scan_date": results.get("scan_start", ""),
                "duration_seconds": results.get("duration_seconds", 0),
            },
            "summary": {
                "subdomains": len(tr_data.get("subdomains", [])),
                "live_hosts": len(tr_data.get("live_hosts", [])),
                "endpoints": len(tr_data.get("endpoints", [])),
                "vulnerabilities": len([v for v in tr_data.get("vulnerabilities", []) if v.get("vulnerable")]),
                "wstg_status": results.get("checklist_summary", {}).get("status_counts", {}),
            },
            "vulnerabilities": [
                v for v in tr_data.get("vulnerabilities", []) if v.get("vulnerable")
            ],
            "subdomains": tr_data.get("subdomains", []),
            "live_hosts": tr_data.get("live_hosts", []),
            "js_findings": tr_data.get("js_findings", []),
            "checklist": [
                {
                    "id": item.get("id"),
                    "category": item.get("category"),
                    "name": item.get("name"),
                    "status": item.get("status").value if hasattr(item.get("status"), "value") else str(item.get("status")),
                    "severity": item.get("severity"),
                    "urls": item.get("urls", []),
                    "evidence": item.get("evidence", []),
                    "recommendations": item.get("recommendations", ""),
                }
                for item in results.get("checklist", [])
            ],
        }

        with open(path, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, ensure_ascii=False, default=str)
        self.logger.info(f"JSON report generated: {path}")

    # ──────────────────────────────────────────────────────────────────
    # MARKDOWN REPORT
    # ──────────────────────────────────────────────────────────────────
    def _generate_markdown(self, results: Dict, path: str):
        """Generate GitHub-ready Markdown report."""
        target = results.get("targets", ["Unknown"])[0]
        scan_date = results.get("scan_start", "")[:10]
        tr_data = results.get("targets_results", [{}])[0] if results.get("targets_results") else {}
        vulns = [v for v in tr_data.get("vulnerabilities", []) if v.get("vulnerable")]
        sc = results.get("checklist_summary", {}).get("status_counts", {})

        md = f"""# 🔒 Security Assessment Report

> **OWASP WSTG Bug Bounty Framework** | [github.com/scp2801/owaspwstg-framework](https://github.com/scp2801/owaspwstg-framework)

---

## 📋 Scan Information

| Field | Value |
|-------|-------|
| **Target** | `{target}` |
| **Scan Date** | {scan_date} |
| **Scan ID** | {results.get('scan_id', 'N/A')} |
| **Duration** | {results.get('duration_seconds', 0)} seconds |
| **Framework** | OWASP WSTG v4.2 |

---

## 📊 Executive Summary

| Severity | Count |
|----------|-------|
| 🔴 **Critical** | {len([v for v in vulns if v.get('severity') == 'Critical'])} |
| 🟠 **High** | {len([v for v in vulns if v.get('severity') == 'High'])} |
| 🟡 **Medium** | {len([v for v in vulns if v.get('severity') == 'Medium'])} |
| 🔵 **Low** | {len([v for v in vulns if v.get('severity') == 'Low'])} |

### Reconnaissance Results

| Metric | Value |
|--------|-------|
| Subdomains Discovered | {len(tr_data.get('subdomains', []))} |
| Live Hosts | {len(tr_data.get('live_hosts', []))} |
| Endpoints Crawled | {len(tr_data.get('endpoints', []))} |
| Parameters Found | {len(tr_data.get('parameters', []))} |

### WSTG Checklist Status

| Status | Count |
|--------|-------|
| ✅ PASS | {sc.get('PASS', 0)} |
| ❌ FAIL | {sc.get('FAIL', 0)} |
| ⚠️ REVIEW | {sc.get('REVIEW', 0)} |
| ⬜ NOT TESTED | {sc.get('NOT_TESTED', 0)} |

---

## 🐛 Vulnerability Findings

"""
        if vulns:
            for i, v in enumerate(sorted(vulns, key=lambda x: {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}.get(x.get("severity", "Low"), 3)), 1):
                sev_emoji = {"Critical": "🔴", "High": "🟠", "Medium": "🟡", "Low": "🔵", "Info": "⚪"}.get(v.get("severity", "Info"), "⚪")
                md += f"""### {i}. {sev_emoji} {v.get('name', 'Unknown')}

| Field | Value |
|-------|-------|
| **Severity** | {v.get('severity', 'N/A')} |
| **WSTG ID** | {v.get('wstg_id', 'N/A')} |
| **CWE** | {v.get('cwe', 'N/A')} |
| **CVSS** | {v.get('cvss', 'N/A')} |
| **URL** | `{v.get('url', 'N/A')}` |

**Description:** {v.get('description', '')}

**Evidence:**
```
{v.get('evidence', 'No evidence collected')[:300]}
```

**Recommendation:** {v.get('recommendation', '')}

---

"""
        else:
            md += "_No exploitable vulnerabilities found._\n\n---\n\n"

        # WSTG Checklist
        md += "## 📋 OWASP WSTG Checklist\n\n"
        md += "| WSTG ID | Category | Test Name | Status | Severity |\n"
        md += "|---------|----------|-----------|--------|----------|\n"

        for item in results.get("checklist", []):
            status = item.get("status", "NOT_TESTED")
            status_val = status.value if hasattr(status, "value") else str(status)
            status_emoji = {"PASS": "✅", "FAIL": "❌", "REVIEW": "⚠️", "NOT_TESTED": "⬜"}.get(status_val, "⬜")
            md += f"| `{item.get('id', '')}` | {item.get('category', '')} | {item.get('name', '')} | {status_emoji} {status_val} | {item.get('severity', '')} |\n"

        md += "\n---\n\n"
        md += "> ⚠️ **Legal Disclaimer:** This report is for authorized security testing only. All findings must be responsibly disclosed.\n"

        with open(path, "w", encoding="utf-8") as f:
            f.write(md)
        self.logger.info(f"Markdown report generated: {path}")

    # ──────────────────────────────────────────────────────────────────
    # TXT REPORT
    # ──────────────────────────────────────────────────────────────────
    def _generate_txt(self, results: Dict, path: str):
        """Generate plain text report."""
        target = results.get("targets", ["Unknown"])[0]
        scan_date = results.get("scan_start", "")[:10]
        tr_data = results.get("targets_results", [{}])[0] if results.get("targets_results") else {}
        vulns = [v for v in tr_data.get("vulnerabilities", []) if v.get("vulnerable")]
        sc = results.get("checklist_summary", {}).get("status_counts", {})

        lines = [
            "=" * 70,
            "OWASP WSTG BUG BOUNTY FRAMEWORK - SECURITY ASSESSMENT REPORT",
            "=" * 70,
            f"Target:    {target}",
            f"Date:      {scan_date}",
            f"Scan ID:   {results.get('scan_id', 'N/A')}",
            f"Duration:  {results.get('duration_seconds', 0)} seconds",
            "",
            "=" * 70,
            "EXECUTIVE SUMMARY",
            "=" * 70,
            f"Subdomains:   {len(tr_data.get('subdomains', []))}",
            f"Live Hosts:   {len(tr_data.get('live_hosts', []))}",
            f"Endpoints:    {len(tr_data.get('endpoints', []))}",
            f"Vulns Found:  {len(vulns)}",
            "",
            f"WSTG PASS:     {sc.get('PASS', 0)}",
            f"WSTG FAIL:     {sc.get('FAIL', 0)}",
            f"WSTG REVIEW:   {sc.get('REVIEW', 0)}",
            f"NOT TESTED:    {sc.get('NOT_TESTED', 0)}",
            "",
            "=" * 70,
            "VULNERABILITY FINDINGS",
            "=" * 70,
        ]

        if vulns:
            for i, v in enumerate(vulns, 1):
                lines += [
                    f"\n[{i}] {v.get('name', 'Unknown')}",
                    f"    Severity:  {v.get('severity', 'N/A')}",
                    f"    WSTG ID:   {v.get('wstg_id', 'N/A')}",
                    f"    CWE:       {v.get('cwe', 'N/A')}",
                    f"    CVSS:      {v.get('cvss', 'N/A')}",
                    f"    URL:       {v.get('url', 'N/A')}",
                    f"    Evidence:  {v.get('evidence', '')[:150]}",
                    f"    Fix:       {v.get('recommendation', '')[:150]}",
                ]
        else:
            lines.append("No exploitable vulnerabilities found.")

        lines += [
            "",
            "=" * 70,
            "OWASP WSTG CHECKLIST",
            "=" * 70,
        ]

        for item in results.get("checklist", []):
            status = item.get("status", "NOT_TESTED")
            status_val = status.value if hasattr(status, "value") else str(status)
            lines.append(
                f"{item.get('id', ''):20} | {status_val:12} | {item.get('severity', ''):8} | {item.get('name', '')[:40]}"
            )

        lines += [
            "",
            "=" * 70,
            "DISCLAIMER: For authorized security testing only.",
            "Generated by OWASP WSTG Framework - github.com/scp2801/owaspwstg-framework",
            "=" * 70,
        ]

        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        self.logger.info(f"TXT report generated: {path}")
